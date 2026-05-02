import openpyxl

try:
    wb = openpyxl.load_workbook('ligas_automatico_v5 (4).xlsx', data_only=True)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            for cell in row:
                if cell and (cell == 659691 or cell == 1963 or str(cell) == '659691' or str(cell) == '1963'):
                    print(f"Found ID in sheet: {sheetname}")
                    print(f"  Row: {row}")
except Exception as e:
    print("Error:", e)
