import openpyxl

try:
    wb = openpyxl.load_workbook('ligas_automatico_v5 (4).xlsx', data_only=True)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            for cell in row:
                if cell and isinstance(cell, str) and 'flashscore.com' in cell:
                    print(f"Found URL in sheet: {sheetname}")
                    print(f"  Row: {row}")
except Exception as e:
    print("Error:", e)
