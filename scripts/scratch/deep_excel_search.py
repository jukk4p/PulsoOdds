import openpyxl

try:
    wb = openpyxl.load_workbook('ligas_automatico_v5 (4).xlsx', data_only=True)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str) and ('MLS' in cell or 'Brasil' in cell):
                    print(f"Found in sheet: {sheetname}")
                    print(f"  Row: {row}")
                    # Stop after first few to not flood output
                    break
            else: continue
            break
except Exception as e:
    print("Error:", e)
