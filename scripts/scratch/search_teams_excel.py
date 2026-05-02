import openpyxl

try:
    wb = openpyxl.load_workbook('ligas_automatico_v5 (4).xlsx', data_only=True)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
            for c_idx, cell in enumerate(row):
                if cell and isinstance(cell, str) and ('Miami' in cell or 'Palmeiras' in cell):
                    print(f"Found '{cell}' in sheet '{sheetname}' at row {r_idx+1}")
except Exception as e:
    print("Error:", e)
